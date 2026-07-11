import os
import json
import re
import smtplib
from io import BytesIO
from datetime import date, datetime
from email.message import EmailMessage
from urllib.parse import parse_qs, urlparse
import unicodedata
from functools import wraps

from flask import Flask, jsonify, request, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import UniqueConstraint, text, inspect
import requests

# ==========================================
# 1. CONFIGURAÇÕES E SETUP
# ==========================================
app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-this')

basedir = os.path.abspath(os.path.dirname(__file__))

# --- NOVO CÓDIGO DO BANCO DE DADOS ---
db_url = os.getenv('DATABASE_URL', 'sqlite:///' + os.path.join(basedir, 'micro_banco.db'))
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
produtos_cache = {}

EMPRESAS_DISPONIVEIS = ['Varejo', 'Redemac', 'Alternativa', 'Granteck', 'Especial']

SMTP_HOST = os.getenv('SMTP_HOST', 'smtp.gmail.com').strip()
SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
SMTP_USER = os.getenv('SMTP_USER', 'ecalliari@gmail.com').strip()
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD', 'ahtl licw snuv tfzw')
SMTP_USE_TLS = os.getenv('SMTP_USE_TLS', 'true').strip().lower() in ('1', 'true', 'yes', 'on')


# ==========================================
# 2. MODELOS DE BANCO DE DADOS
# ==========================================
class Usuario(db.Model):
    __table_args__ = (UniqueConstraint('email', 'cnpj', name='uq_usuario_email_cnpj'),)

    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(80), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    cnpj = db.Column(db.String(14), nullable=False)
    celular = db.Column(db.String(20), unique=True, nullable=True)
    associacoes = db.Column(db.String(255), nullable=False, default='Varejo')
    empresa_ativa = db.Column(db.String(120), nullable=False, default='Varejo')
    perm_download_tabelas = db.Column(db.Boolean, nullable=False, default=True)
    perm_fazer_pedido = db.Column(db.Boolean, nullable=False, default=True)
    perm_solicitar_visita = db.Column(db.Boolean, nullable=False, default=True)
    perm_visualizar_pedidos = db.Column(db.Boolean, nullable=False, default=True)
    is_admin = db.Column(db.Boolean, nullable=False, default=False)
    senha = db.Column(db.String(120), nullable=True)

class Agendamento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    data_hora = db.Column(db.DateTime, nullable=False)
    observacao = db.Column(db.String(255), nullable=True)

class Pedido(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    data_pedido = db.Column(db.DateTime, default=datetime.now, nullable=False)
    valor_total = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(50), default='Pendente', nullable=False)
    itens = db.relationship('ItensPedido', backref='pedido', lazy=True, cascade='all, delete-orphan')

class ItensPedido(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    pedido_id = db.Column(db.Integer, db.ForeignKey('pedido.id'), nullable=False)
    referencia_produto = db.Column(db.String(120), nullable=False)
    quantidade = db.Column(db.Integer, nullable=False)
    valor_unitario = db.Column(db.Float, nullable=False)
    valor_total = db.Column(db.Float, nullable=False)

class DeletarUsuarios(db.Model):
    __tablename__ = 'deletar_usuarios'
    id = db.Column(db.Integer, primary_key=True)

    @staticmethod
    def deletar_todos():
        num_rows = db.session.query(Usuario).delete()
        db.session.commit()
        return num_rows


# ==========================================
# 3. FUNÇÕES AUXILIARES E ÚTEIS
# ==========================================
def get_payload():
    return request.get_json(silent=True) or request.form

def normalizar_cnpj(cnpj):
    return re.sub(r'\D', '', str(cnpj or ''))

def parse_empresas(valor_empresas):
    if isinstance(valor_empresas, list):
        candidatos = valor_empresas
    else:
        texto = str(valor_empresas or '')
        candidatos = [item.strip() for item in texto.replace(';', ',').split(',') if item.strip()]

    filtradas = [e for e in candidatos if e in EMPRESAS_DISPONIVEIS]
    return list(dict.fromkeys(filtradas)) or ['Varejo']

def obter_permissoes_usuario(usuario):
    if usuario.is_admin:
        return {k: True for k in ['download_tabelas', 'fazer_pedido', 'solicitar_visita', 'visualizar_pedidos', 'atualizar_tabelas']}
    return {
        'download_tabelas': bool(usuario.perm_download_tabelas),
        'fazer_pedido': bool(usuario.perm_fazer_pedido),
        'solicitar_visita': bool(usuario.perm_solicitar_visita),
        'visualizar_pedidos': bool(usuario.perm_visualizar_pedidos),
        'atualizar_tabelas': False,
    }

def obter_arquivo_tabela(associacao):
    arquivo_map = {e: f'static/tabelas/tabela_{e.lower()}.xlsx' for e in EMPRESAS_DISPONIVEIS}
    return arquivo_map.get(associacao, 'static/tabelas/tabela.xlsx')

def normalizar_texto_coluna(valor):
    texto = str(valor or '').strip().lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('utf-8')
    return re.sub(r'[^a-z0-9]+', '', texto)

def encontrar_indice_coluna(cabecalho, aliases):
    aliases_norm = {normalizar_texto_coluna(a) for a in aliases}
    for idx, coluna in enumerate(cabecalho):
        if normalizar_texto_coluna(coluna) in aliases_norm:
            return idx
    return None

def normalizar_decimal_generico(valor):
    if valor is None: return None
    texto = re.sub(r'[^0-9,.-]', '', str(valor).strip())
    if not texto: return None
    if ',' in texto and '.' in texto:
        texto = texto.replace('.', '').replace(',', '.')
    else:
        texto = texto.replace(',', '.')
    try:
        return float(texto)
    except ValueError:
        return None

def normalizar_data_generica(valor):
    if not valor: return None
    if isinstance(valor, datetime): return valor.date()
    if isinstance(valor, date): return valor
    
    if isinstance(valor, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            convertido = from_excel(valor)
            return convertido.date() if isinstance(convertido, datetime) else convertido
        except Exception:
            return None

    texto = str(valor).strip().split()[0]
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None

def geocodificar_endereco(endereco):
    if not endereco: return None, None
    try:
        resp = requests.get(
            'https://nominatim.openstreetmap.org/search',
            params={'q': endereco, 'format': 'json', 'limit': 1, 'countrycodes': 'br'},
            headers={'User-Agent': 'app-vendas/1.0 (sales-map geocoding)'},
            timeout=15
        )
        dados = resp.json() if resp.status_code == 200 else []
        if dados:
            return normalizar_decimal_generico(dados[0].get('lat')), normalizar_decimal_generico(dados[0].get('lon'))
    except requests.RequestException:
        pass
    return None, None


def montar_endereco_cliente(endereco, bairro, cidade, estado, cep):
    partes = [
        str(endereco or '').strip(),
        str(bairro or '').strip(),
        str(cidade or '').strip(),
        str(estado or '').strip(),
        str(cep or '').strip(),
        'Brasil',
    ]
    return ', '.join([parte for parte in partes if parte])

def serializar_usuario(usuario):
    empresas = parse_empresas(usuario.associacoes)
    return {
        'id': usuario.id,
        'nome': usuario.nome,
        'email': usuario.email,
        'cnpj': usuario.cnpj,
        'celular': usuario.celular,
        'empresas': empresas,
        'empresa_ativa': usuario.empresa_ativa if usuario.empresa_ativa in empresas else empresas[0],
        'permissoes': obter_permissoes_usuario(usuario),
        'is_admin': bool(usuario.is_admin),
    }

def enviar_email_pedido_admin(usuario, pedido, itens):
    if not usuario.email or not SMTP_HOST: return
    linhas_itens = [
        f"- Ref: {i.get('referencia_produto')} | Qtd: {i.get('quantidade')} | "
        f"Vlr Unit.: R$ {float(i.get('valor_unitario', 0)):.2f} | Vlr Total: R$ {float(i.get('valor_total', 0)):.2f}"
        for i in itens
    ]
    corpo = (
        f"Seu pedido foi finalizado com sucesso na plataforma.\n\n"
        f"Pedido: #{pedido.id}\nData: {pedido.data_pedido.strftime('%d/%m/%Y %H:%M:%S')}\n"
        f"Valor total: R$ {pedido.valor_total:.2f}\n\nDados do cliente:\n"
        f"Nome: {usuario.nome}\nEmail: {usuario.email}\nCNPJ: {usuario.cnpj}\n"
        f"Celular: {usuario.celular or '-'}\nEmpresa ativa: {usuario.empresa_ativa}\n\n"
        f"Itens do pedido:\n{chr(10).join(linhas_itens)}\n"
    )
    msg = EmailMessage()
    msg['Subject'] = f'Confirmacao do pedido #{pedido.id} - {usuario.nome}'
    msg['From'] = SMTP_USER or 'no-reply@sistema.local'
    msg['To'] = usuario.email
    msg.set_content(corpo)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as servidor:
        if SMTP_USE_TLS: servidor.starttls()
        if SMTP_USER and SMTP_PASSWORD: servidor.login(SMTP_USER, SMTP_PASSWORD)
        servidor.send_message(msg)


# ==========================================
# 4. INICIALIZAÇÃO DO BANCO
# ==========================================
def coluna_existe(tabela, coluna):
    inspector = inspect(db.engine)
    # Verifica se a tabela existe antes de procurar a coluna
    if not inspector.has_table(tabela):
        return False
    # Busca os nomes de todas as colunas da tabela
    colunas = [coluna_info['name'] for coluna_info in inspector.get_columns(tabela)]
    return coluna in colunas

def atualizar_schema_legado():
    cols_add = {
        'cnpj': 'VARCHAR(14)', 'associacoes': "VARCHAR(255) DEFAULT 'Varejo'",
        'empresa_ativa': "VARCHAR(120) DEFAULT 'Varejo'", 
        'perm_download_tabelas': 'BOOLEAN DEFAULT TRUE',
        'perm_fazer_pedido': 'BOOLEAN DEFAULT TRUE', 
        'perm_solicitar_visita': 'BOOLEAN DEFAULT TRUE',
        'perm_visualizar_pedidos': 'BOOLEAN DEFAULT TRUE', 
        'is_admin': 'BOOLEAN DEFAULT FALSE'
    }
    alterado = False
    for col, tip in cols_add.items():
        if not coluna_existe('usuario', col):
            db.session.execute(text(f'ALTER TABLE usuario ADD COLUMN {col} {tip}'))
            alterado = True
    if alterado: db.session.commit()

def garantir_usuario_admin():
    email = os.getenv('ADMIN_EMAIL', 'admin@sistema.local').strip().lower()
    cnpj = normalizar_cnpj(os.getenv('ADMIN_CNPJ', '00000000000000'))
    admin = Usuario.query.filter_by(email=email, cnpj=cnpj).first()
    
    if not admin:
        admin = Usuario(nome='Administrador', email=email, cnpj=cnpj, celular='(00)00000-0000',
                        associacoes=','.join(EMPRESAS_DISPONIVEIS), empresa_ativa=EMPRESAS_DISPONIVEIS[0],
                        senha=os.getenv('ADMIN_SENHA', 'admin123'), is_admin=True)
        db.session.add(admin)
    else:
        admin.is_admin = True
    db.session.commit()

with app.app_context():
    db.create_all()
    atualizar_schema_legado()
    garantir_usuario_admin()


# ==========================================
# 5. DECORADORES
# ==========================================
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_data' not in session:
            return jsonify({'erro': 'Usuário não autenticado'}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user_data = session.get('user_data', {})
        if not user_data.get('is_admin'):
            return jsonify({'erro': 'Acesso restrito para administradores.'}), 403
        return f(*args, **kwargs)
    return decorated_function


# ==========================================
# 6. ROTAS / ENDPOINTS
# ==========================================
@app.route('/adicionar', methods=['POST'])
def adicionar_usuario():
    payload = get_payload()
    nome, email, celular, senha = [payload.get(k, '').strip() for k in ('nome', 'email', 'celular', 'senha')]
    cnpj = normalizar_cnpj(payload.get('cnpj'))
    email = email.lower()

    if not all([nome, email, celular, cnpj, senha]):
        return jsonify({'erro': 'Todos os campos são obrigatórios.'}), 400
    if len(cnpj) != 14:
        return jsonify({'erro': 'CNPJ inválido. Informe 14 dígitos.'}), 400

    if Usuario.query.filter_by(email=email, cnpj=cnpj).first():
        return jsonify({'erro': 'Já existe usuário com este email e CNPJ.'}), 409
    if Usuario.query.filter_by(celular=celular).first():
        return jsonify({'erro': 'Já existe usuário com este celular.'}), 409

    empresas_lista = parse_empresas(payload.get('empresas', payload.get('associacao', 'Varejo')))
    novo_usuario = Usuario(
        nome=nome, email=email, cnpj=cnpj, celular=celular, senha=senha,
        associacoes=','.join(empresas_lista), empresa_ativa=empresas_lista[0],
        perm_download_tabelas=str(payload.get('perm_download_tabelas', 'true')).lower() in ('true', '1'),
        perm_fazer_pedido=str(payload.get('perm_fazer_pedido', 'true')).lower() in ('true', '1'),
        perm_solicitar_visita=str(payload.get('perm_solicitar_visita', 'true')).lower() in ('true', '1'),
        perm_visualizar_pedidos=str(payload.get('perm_visualizar_pedidos', 'true')).lower() in ('true', '1')
    )
    db.session.add(novo_usuario)
    db.session.commit()

    session['user_data'] = serializar_usuario(novo_usuario)
    return jsonify(session['user_data']), 201

@app.route('/login', methods=['POST'])
def login():
    payload = get_payload()
    email = (payload.get('email') or '').strip().lower()
    cnpj = normalizar_cnpj(payload.get('cnpj'))
    senha = payload.get('senha')

    if not all([email, cnpj, senha]):
        return jsonify({'erro': 'Email, CNPJ e senha são obrigatórios.'}), 400

    usuario = Usuario.query.filter_by(email=email, cnpj=cnpj).first()
    if not usuario:
        return jsonify({'erro': 'Usuário não encontrado.'}), 404
    if usuario.senha != senha:
        return jsonify({'erro': 'Senha incorreta.'}), 401
    
    session['user_data'] = serializar_usuario(usuario)
    return jsonify(session['user_data']), 200

@app.route('/selecionar_empresa', methods=['POST'])
def selecionar_empresa():
    payload = get_payload()
    usuario_id, empresa = payload.get('usuario_id'), (payload.get('empresa') or '').strip()

    if not usuario_id or not empresa:
        return jsonify({'erro': 'usuario_id e empresa são obrigatórios.'}), 400

    usuario = Usuario.query.get(usuario_id)
    if not usuario: return jsonify({'erro': 'Usuário não encontrado.'}), 404

    if empresa not in parse_empresas(usuario.associacoes):
        return jsonify({'erro': 'Acesso negado à empresa selecionada.'}), 403

    usuario.empresa_ativa = empresa
    db.session.commit()
    session['user_data'] = serializar_usuario(usuario)
    return jsonify(session['user_data']), 200

@app.route('/adicionaragenda', methods=['POST'])
def adicionaragenda():
    payload = get_payload()
    if not payload.get('usuario_id') or not payload.get('data_hora'):
        return jsonify({'erro': 'usuario_id e data_hora são obrigatórios.'}), 400
    try:
        data_hora = datetime.fromisoformat(str(payload.get('data_hora')))
    except ValueError:
        return jsonify({'erro': 'Formato inválido. Use ISO.'}), 400

    novo_agendamento = Agendamento(usuario_id=payload.get('usuario_id'), data_hora=data_hora, observacao=payload.get('observacao', ''))
    db.session.add(novo_agendamento)
    db.session.commit()
    return jsonify({'mensagem': 'Agendamento realizado!'}), 201

@app.route('/usuarios', methods=['GET'])
def listar_usuarios():
    return jsonify([serializar_usuario(u) for u in Usuario.query.all()])

@app.route('/carregar_produtos', methods=['GET'])
def carregar_produtos():
    try:
        from openpyxl import load_workbook
        
        associacao = request.args.get('associacao') or request.args.get('empresa') or session.get('user_data', {}).get('empresa_ativa', 'Varejo')
        arquivo_caminho = os.path.join(basedir, obter_arquivo_tabela(associacao))
        
        if not os.path.exists(arquivo_caminho):
            return jsonify({'erro': f'Arquivo não encontrado para {associacao}'}), 404

        mtime = os.path.getmtime(arquivo_caminho)
        if produtos_cache.get(associacao, {}).get('mtime') == mtime:
            return jsonify(produtos_cache[associacao]['produtos']), 200

        wb = load_workbook(filename=arquivo_caminho, data_only=True, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        if not rows: return jsonify({'erro': 'Arquivo vazio'}), 400

        cabecalho = [str(c or '').strip().lower() for c in rows[0]]
        idxs = {k: encontrar_indice_coluna(cabecalho, [k]) for k in ['referência', 'descrição', 'emb.', 'preço final']}

        produtos = []
        for r in rows[1:]:
            try:
                ref = str(r[idxs['referência']]).strip()
                desc = str(r[idxs['descrição']]).strip() if idxs['descrição'] is not None else ''
                emb = normalizar_decimal_generico(r[idxs['emb.']])
                preco = normalizar_decimal_generico(r[idxs['preço final']])
                if ref and emb and preco:
                    produtos.append({'referencia': ref, 'descricao': desc, 'quantidade_embalagem': round(emb, 2), 'valor': round(preco, 2)})
            except IndexError: continue

        produtos_cache[associacao] = {'mtime': mtime, 'produtos': produtos}
        return jsonify(produtos), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/atualizar_tabelas', methods=['POST'])
def atualizar_tabelas():
    usuario_id = request.form.get('usuario_id', type=int)
    usuario = Usuario.query.get(usuario_id)
    if not usuario or not usuario.is_admin:
        return jsonify({'erro': 'Acesso negado. Apenas admins.'}), 403

    arquivo = request.files.get('arquivo_tabela')
    if not arquivo or not arquivo.filename.lower().endswith('.xlsx'):
        return jsonify({'erro': 'Envie um arquivo .xlsx válido.'}), 400

    associacao = request.form.get('associacao') or usuario.empresa_ativa or 'Varejo'
    destino = os.path.join(basedir, obter_arquivo_tabela(associacao))

    if os.path.exists(destino):
        base, ext = os.path.splitext(destino)
        backup_path = f"{base}_{datetime.now().strftime('%Y%m%d')}{ext}"
        os.rename(destino, backup_path)

    arquivo.save(destino)
    produtos_cache.pop(associacao, None)
    return jsonify({'mensagem': f'Tabela atualizada para {associacao}.'}), 200

@app.route('/salvar_pedido', methods=['POST'])
@login_required
def salvar_pedido():
    dados = get_payload()
    usuario_id = session['user_data']['id']
    usuario = Usuario.query.get(usuario_id)

    if not usuario.is_admin and not usuario.perm_fazer_pedido:
        return jsonify({'erro': 'Sem permissão para fazer pedidos'}), 403

    itens = dados.get('itens', [])
    if not itens: return jsonify({'erro': 'Nenhum item no pedido'}), 400

    try:
        valor_total = sum(i['valor_total'] for i in itens)
        novo_pedido = Pedido(usuario_id=usuario_id, valor_total=valor_total)
        db.session.add(novo_pedido)
        db.session.flush()

        for item in itens:
            db.session.add(ItensPedido(
                pedido_id=novo_pedido.id, referencia_produto=item['referencia_produto'],
                quantidade=item['quantidade'], valor_unitario=item['valor_unitario'], valor_total=item['valor_total']
            ))
        db.session.commit()

        aviso = None
        try: enviar_email_pedido_admin(usuario, novo_pedido, itens)
        except Exception as e: aviso = str(e)

        return jsonify({'mensagem': 'Pedido salvo!', 'pedido_id': novo_pedido.id, 'aviso_email': aviso}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 500

@app.route('/meus_pedidos', methods=['GET'])
@login_required
def listar_meus_pedidos():
    pedidos = Pedido.query.filter_by(usuario_id=session['user_data']['id']).order_by(Pedido.data_pedido.desc()).all()
    return jsonify([{
        'id': p.id, 'data_pedido': p.data_pedido.isoformat(),
        'valor_total': round(p.valor_total, 2), 'status': p.status, 'quantidade_itens': len(p.itens)
    } for p in pedidos]), 200

@app.route('/meus_pedidos/<int:pedido_id>/itens', methods=['GET'])
@login_required
def listar_meus_itens_pedido(pedido_id):
    pedido = Pedido.query.filter_by(id=pedido_id, usuario_id=session['user_data']['id']).first()
    if not pedido: return jsonify({'erro': 'Pedido não encontrado'}), 404

    return jsonify({
        'pedido_id': pedido.id, 'valor_total': round(pedido.valor_total, 2),
        'itens': [{'id': i.id, 'referencia_produto': i.referencia_produto, 'quantidade': i.quantidade,
                   'valor_unitario': round(i.valor_unitario, 2), 'valor_total': round(i.valor_total, 2)} for i in pedido.itens]
    }), 200

@app.route('/pedidos_usuario/<int:usuario_id>', methods=['GET'])
def listar_pedidos_usuario(usuario_id):
    # Rota mantida por compatibilidade (considerar mesclar lógica com /meus_pedidos no front futuramente)
    pedidos = Pedido.query.filter_by(usuario_id=usuario_id).order_by(Pedido.data_pedido.desc()).all()
    return jsonify([{'id': p.id, 'valor_total': round(p.valor_total, 2), 'status': p.status, 'quantidade_itens': len(p.itens)} for p in pedidos]), 200

@app.route('/deletar', methods=['POST'])
def deletar():
    return f"{DeletarUsuarios.deletar_todos()} usuários deletados do micro banco!"


@app.route('/dados_mapa_vendas', methods=['GET'])
def dados_mapa_vendas():
    try:
        try:
            from openpyxl import load_workbook
        except Exception as erro_openpyxl:
            return jsonify({'erro': f'Falha ao carregar openpyxl: {str(erro_openpyxl)}'}), 500

        usuario_id = request.args.get('usuario_id', type=int)
        data_inicio_raw = (request.args.get('data_inicio') or '').strip()
        data_fim_raw = (request.args.get('data_fim') or '').strip()

        if not usuario_id:
            return jsonify({'erro': 'usuario_id é obrigatório.'}), 400

        data_inicio = None
        data_fim = None
        if data_inicio_raw:
            try:
                data_inicio = datetime.strptime(data_inicio_raw, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'erro': 'data_inicio inválida. Use YYYY-MM-DD.'}), 400

        if data_fim_raw:
            try:
                data_fim = datetime.strptime(data_fim_raw, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'erro': 'data_fim inválida. Use YYYY-MM-DD.'}), 400

        if data_inicio and data_fim and data_inicio > data_fim:
            return jsonify({'erro': 'data_inicio não pode ser maior que data_fim.'}), 400

        usuario = Usuario.query.get(usuario_id)
        if not usuario or not usuario.is_admin:
            return jsonify({'erro': 'Apenas administrador pode acessar o mapa de vendas.'}), 403

        arquivo_vendas = os.path.join(basedir, 'static', 'vendas', 'vendas_feitas.xlsx')
        arquivo_carteira = os.path.join(basedir, 'static', 'carteira', 'carteira_clientes.xlsx')

        if not os.path.exists(arquivo_vendas) or not os.path.exists(arquivo_carteira):
            return jsonify({'erro': 'Arquivos de vendas ou carteira não encontrados.'}), 404

        # --- PROCESSA VENDAS ---
        wb_vendas = load_workbook(filename=arquivo_vendas, data_only=True, read_only=True)
        ws_vendas = wb_vendas.active
        linhas_vendas = list(ws_vendas.iter_rows(values_only=True))
        if not linhas_vendas:
            return jsonify({'erro': 'Planilha de vendas está vazia.'}), 400

        cabecalho_vendas = [str(c or '').strip() for c in linhas_vendas[0]]
        idx_codigo_venda = encontrar_indice_coluna(cabecalho_vendas, ['Codigo', 'Cdigo'])
        idx_nome_venda = encontrar_indice_coluna(cabecalho_vendas, ['Nome Cliente'])
        idx_valor_venda = encontrar_indice_coluna(cabecalho_vendas, ['Valor Produtos'])
        idx_data_emissao = encontrar_indice_coluna(cabecalho_vendas, ['Data Emissao', 'Data Emissão'])

        if idx_nome_venda is None or idx_valor_venda is None:
            return jsonify({'erro': 'Colunas obrigatórias faltando em vendas_feitas.xlsx.'}), 400

        vendas_por_cliente = {}
        for linha in linhas_vendas[1:]:
            nome_cliente = str(linha[idx_nome_venda] or '').strip() if idx_nome_venda < len(linha) else ''
            if not nome_cliente: continue

            if data_inicio or data_fim:
                data_emissao = normalizar_data_generica(linha[idx_data_emissao] if idx_data_emissao < len(linha) else None)
                if data_emissao is None: continue
                if data_inicio and data_emissao < data_inicio: continue
                if data_fim and data_emissao > data_fim: continue

            codigo = str(linha[idx_codigo_venda] or '').strip() if idx_codigo_venda is not None and idx_codigo_venda < len(linha) else ''
            valor = normalizar_decimal_generico(linha[idx_valor_venda] if idx_valor_venda < len(linha) else None) or 0.0

            chave = normalizar_texto_coluna(nome_cliente)
            if chave not in vendas_por_cliente:
                vendas_por_cliente[chave] = {'codigo': codigo, 'nome_cliente': nome_cliente, 'valor_produtos': 0.0}
            vendas_por_cliente[chave]['valor_produtos'] += valor

        # --- PROCESSA CARTEIRA ---
        wb_carteira = load_workbook(filename=arquivo_carteira, data_only=True, read_only=True)
        ws_carteira = wb_carteira.active
        linhas_carteira = list(ws_carteira.iter_rows(values_only=True))

        cabecalho_carteira = [str(c or '').strip() for c in linhas_carteira[0]]
        idxs = {k: encontrar_indice_coluna(cabecalho_carteira, v) for k, v in {
            'codigo': ['Codigo', 'Cdigo'], 'nome': ['Nome Cliente'], 'endereco': ['Endereco', 'Endereo'],
            'bairro': ['Bairro'], 'cidade': ['Cidade'], 'estado': ['Estado', 'UF'], 'cep': ['CEP'],
            'lat': ['Latitude', 'Lat'], 'lon': ['Longitude', 'Lon']
        }.items()}

        cache_geo_path = os.path.join(basedir, 'static', 'carteira', 'geocode_cache.json')
        geocode_cache = {}
        if os.path.exists(cache_geo_path):
            try:
                with open(cache_geo_path, 'r', encoding='utf-8') as f: geocode_cache = json.load(f)
            except Exception: pass

        pontos, cache_alterado, clientes_sem_coord = [], False, 0

        for linha in linhas_carteira[1:]:
            nome = str(linha[idxs['nome']] or '').strip() if idxs['nome'] is not None and idxs['nome'] < len(linha) else ''
            if not nome: continue

            end = linha[idxs['endereco']] if idxs['endereco'] is not None and idxs['endereco'] < len(linha) else ''
            bairro = linha[idxs['bairro']] if idxs['bairro'] is not None and idxs['bairro'] < len(linha) else ''
            cidade = linha[idxs['cidade']] if idxs['cidade'] is not None and idxs['cidade'] < len(linha) else ''
            estado = linha[idxs['estado']] if idxs['estado'] is not None and idxs['estado'] < len(linha) else ''
            cep = linha[idxs['cep']] if idxs['cep'] is not None and idxs['cep'] < len(linha) else ''
            end_completo = montar_endereco_cliente(end, bairro, cidade, estado, cep)

            lat = normalizar_decimal_generico(linha[idxs['lat']]) if idxs['lat'] is not None and idxs['lat'] < len(linha) else None
            lon = normalizar_decimal_generico(linha[idxs['lon']]) if idxs['lon'] is not None and idxs['lon'] < len(linha) else None

            cache_key = f"{normalizar_texto_coluna(nome)}|{normalizar_texto_coluna(end_completo)}"
            if (lat is None or lon is None) and cache_key in geocode_cache:
                lat, lon = geocode_cache[cache_key].get('lat'), geocode_cache[cache_key].get('lon')

            if (lat is None or lon is None) and end_completo:
                lat, lon = geocodificar_endereco(end_completo)
                if lat is not None and lon is not None:
                    geocode_cache[cache_key] = {'lat': lat, 'lon': lon}
                    cache_alterado = True

            if lat is None or lon is None:
                clientes_sem_coord += 1
                continue

            chave_nome = normalizar_texto_coluna(nome)
            venda = vendas_por_cliente.get(chave_nome)
            pontos.append({
                'codigo': str(linha[idxs['codigo']] or '').strip() if idxs['codigo'] is not None and idxs['codigo'] < len(linha) else '',
                'nome_cliente': nome,
                'valor_produtos': round(float(venda['valor_produtos']), 2) if venda else 0.0,
                'endereco': end_completo, 'latitude': lat, 'longitude': lon,
                'fez_venda': bool(venda), 'cor': 'green' if venda else 'red',
            })

        if cache_alterado:
            try:
                with open(cache_geo_path, 'w', encoding='utf-8') as f: json.dump(geocode_cache, f, ensure_ascii=False)
            except Exception: pass

        return jsonify({
            'resumo': {
                'total_clientes_carteira': len(linhas_carteira) - 1,
                'clientes_com_venda': sum(1 for p in pontos if p['fez_venda']),
                'clientes_sem_venda': sum(1 for p in pontos if not p['fez_venda']),
                'clientes_sem_coordenada': clientes_sem_coord,
                'periodo': {'data_inicio': data_inicio.isoformat() if data_inicio else None, 'data_fim': data_fim.isoformat() if data_fim else None}
            },
            'pontos': pontos,
        }), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)